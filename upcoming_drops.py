import logging
from configparser import ConfigParser
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, DEFAULT_FONT
from how_rare_is_connector import HowRareIs
from openpyxl.utils import get_column_letter
import os


class UpcomingDrops:
    _STYLE_YELLOW_FILL = PatternFill(start_color="F1C232", fill_type="solid")
    _STYLE_BLUE_FILL = PatternFill(start_color="1c4587", fill_type="solid")

    _FONT_BOLD_TITLE = Font(name="Arial", size=14, bold=True)
    _FONT_BOLD_HEADING = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    _FONT_BOLD_HEADING_CHAT = Font(name="Arial", size=18, bold=True)
    _FONT_DATE = Font(name="Arial", size=18, bold=True)
    _FONT_BODY = Font(name="Arial", size=10)

    def __init__(self, filename, warning_title, warning_subtitle, add_sheets_for_days):
        self._drops = HowRareIs()
        self._filename = filename

        self._add_sheets_for_days = add_sheets_for_days

        self._drops_written = 0

        self._warning_title = warning_title
        self._warning_subtitle = warning_subtitle

        self._log = logging.getLogger(__name__)

        self._log_init()

    def _log_init(self):
        self._log.info("Initializing HowRare.IsDropsScraper")
        self._log.info("%s%s %s", " " * 4, "Filename".ljust(30, "."), self._filename)
        self._log.info(
            "%s%s %s", " " * 4, "Warning Title".ljust(30, "."), self._warning_title
        )
        self._log.info(
            "%s%s %s",
            " " * 4,
            "Warning Subtitle".ljust(30, "."),
            self._warning_subtitle,
        )

    def _save_workbook(self, workbook, filename):
        while True:
            try:
                workbook.save(filename)
                return
            except:
                self._log.error(
                    "Unable to export data to file: %s. Do you have Excel open? Please close Excel and press ENTER.",
                    filename,
                )
                input()

    def _draw_warnings(self, worksheet):
        ws = worksheet

        cell_warning_title = ws.cell(row=1, column=1, value=self._warning_title)
        cell_warning_title.alignment = Alignment(horizontal="center")
        cell_warning_title.fill = self._STYLE_YELLOW_FILL
        cell_warning_title.font = self._FONT_BOLD_TITLE

        cell_warning_subtitle = ws.cell(row=2, column=1, value=self._warning_subtitle)
        cell_warning_subtitle.alignment = Alignment(horizontal="center")
        cell_warning_subtitle.fill = self._STYLE_YELLOW_FILL
        cell_warning_subtitle.font = self._FONT_BOLD_TITLE

        ws.merge_cells("A1:I1")
        ws.merge_cells("A2:I2")

    def _as_text(self, value):
        if value is None:
            return ""
        return str(value)

    @property
    def _row_start_data(self):
        return 5 + self._drops_written

    def _draw_headings(self, worksheet):
        ws = worksheet

        headings = [
            "Mint Date",
            "Project Name",
            "EST",
            "UTC",
            "Twitter",
            "Discord",
            "Website",
            "Supply",
            "Mint Price",
        ]

        # Mint Time Heading
        ws.cell(row=3, column=3, value="Mint Time")

        ws.merge_cells("C3:D3")

        # Main Headings
        row = 4
        col = 1

        for heading in headings:
            ws.cell(row=row, column=col, value=heading)

            col += 1

        # Chat Headings
        chat_heading = ws.cell(row=1, column=10, value="Chat's Thoughts")
        chat_heading.fill = self._STYLE_YELLOW_FILL
        chat_heading.font = self._FONT_BOLD_HEADING_CHAT
        chat_heading.alignment = Alignment(horizontal="center")

        ws.column_dimensions[get_column_letter(10)].width = 50
        ws.merge_cells("J1:J2")

        # Style All Heading Cells
        heading_row_start = 3
        heading_row_end = 4
        heading_col_start = 1
        heading_col_end = 11
        for row in range(heading_row_start, heading_row_end + 1):
            for col in range(heading_col_start, heading_col_end):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center")
                cell.font = self._FONT_BOLD_HEADING
                cell.fill = self._STYLE_BLUE_FILL

    def _auto_size_columns(self, worksheet, columns_to_ignore):
        for column_cells in worksheet.columns:
            col_letter = get_column_letter(column_cells[0].column)
            if col_letter not in columns_to_ignore:
                if len(column_cells) > 4:
                    length = (
                        max(len(self._as_text(cell.value)) for cell in column_cells[4:])
                        + 5
                    )
                    worksheet.column_dimensions[col_letter].width = length
                else:
                    self._log.warning(
                        "No data cells to auto-size. Are there any drops today?"
                    )

    def _fonts_to_arial(self, worksheet):
        ws = worksheet
        max_row = ws.max_row

        for row_idx in range(1, max_row):
            for cell in ws[row_idx]:
                cell.font = Font(
                    name="Arial",
                    size=cell.font.size,
                    color=cell.font.color,
                    bold=cell.font.bold,
                )

    def _draw_styling(self, worksheet):
        ws = worksheet

        # Draw Warnings
        self._draw_warnings(ws)

    def _draw_one_day_of_drops(self, worksheet, date, drops):
        ws = worksheet
        drop_count = len(drops)

        # Make Date Cell
        start_row = self._row_start_data
        date_cell = ws.cell(row=start_row, column=1, value=date)
        date_cell.font = self._FONT_DATE
        date_cell.alignment = Alignment(vertical="center", horizontal="center")

        ws.merge_cells(f"A{start_row}:A{start_row + drop_count - 1}")

        for drop in drops:
            row = self._row_start_data
            col = 2

            cell = ws.cell(row=row, column=col, value=drop["project_name"])
            cell.font = self._FONT_BODY
            col += 1

            cell = ws.cell(row=row, column=col, value=drop["time_est"])
            cell.alignment = Alignment(horizontal="center")
            cell.font = self._FONT_BODY
            col += 1

            cell = ws.cell(row=row, column=col, value=drop["time_utc"])
            cell.font = self._FONT_BODY
            cell.alignment = Alignment(horizontal="center")
            col += 1

            cell = ws.cell(
                row=row,
                column=col,
                value='=HYPERLINK("{}", "{}")'.format(
                    drop["twitter_url"], "Twitter Link" if drop["twitter_url"] else None
                ),
            )
            cell.font = self._FONT_BODY
            col_letter = get_column_letter(cell.column)
            worksheet.column_dimensions[col_letter].width = len("Twitter Link")
            col += 1

            cell = ws.cell(
                row=row,
                column=col,
                value='=HYPERLINK("{}", "{}")'.format(
                    drop["discord_url"], "Discord Link" if drop["discord_url"] else None
                ),
            )
            cell.font = self._FONT_BODY
            col_letter = get_column_letter(cell.column)
            worksheet.column_dimensions[col_letter].width = len("Discord Link")
            col += 1

            cell = ws.cell(
                row=row,
                column=col,
                value='=HYPERLINK("{}", "{}")'.format(
                    drop["website_url"], "Website Link" if drop["website_url"] else None
                ),
            )
            cell.font = self._FONT_BODY
            col_letter = get_column_letter(cell.column)
            worksheet.column_dimensions[col_letter].width = len("Website Link")
            col += 1

            cell = ws.cell(row=row, column=col, value=drop["supply"])
            cell.alignment = Alignment(horizontal="center")
            cell.font = self._FONT_BODY
            col += 1

            cell = ws.cell(row=row, column=col, value=drop["mint_price"])
            cell.font = self._FONT_BODY
            cell.alignment = Alignment(horizontal="center")
            col += 1

            self._drops_written += 1

    def create_excel(self, how_many_days):
        drops_workbook = openpyxl.Workbook()
        ws = None

        # Remove default sheet
        drops_workbook.remove(drops_workbook.active)
        DEFAULT_FONT.name = "Arial"

        if not self._add_sheets_for_days:
            # Create nice named sheet
            ws = drops_workbook.create_sheet("Upcoming Drops")

            # Create styling
            self._draw_styling(ws)

            # Draw Headings
            self._draw_headings(ws)

        self._log.info("Acquiring drops...")
        drops = self._drops.get_drops()
        self._log.info("Creating Excel document ...")
        self._log.info(
            "Printing %s of %s days.",
            min(how_many_days, len(drops.keys())),
            len(drops.keys()),
        )
        for i, drop in enumerate(drops):
            if i < how_many_days:
                if self._add_sheets_for_days:
                    self._drops_written = 0
                    ws = drops_workbook.create_sheet(drop.replace("/", "-"))
                    # Create styling
                    self._draw_styling(ws)

                    # Draw Headings
                    self._draw_headings(ws)
                self._draw_one_day_of_drops(ws, drop, drops[drop])

            # Resize Columns
            self._auto_size_columns(ws, ["J", "E", "F", "G"])

        self._save_workbook(drops_workbook, self._filename)
        self._log.info("Drops saved to %s.", self._filename)


def get_default_config():
    return {
        "file_info": {"filename": "UpcomingDrops.xlsx"},
        "appearance": {
            "warning_title": "This is not financial Advice. Do your own research.",
            "warning_subtitle": "Having a project listed on this sheet is not an endorsement of that project.",
        },
        "functionality": {"days_to_export": "1", "additional_days_add_sheets": "False"},
        "debug": {"log_level": "info"},
    }


def create_default_config(filename):
    if not os.path.isfile(filename):
        config = ConfigParser()
        default_config = get_default_config()

        for key in default_config:
            config.add_section(key)
            for subkey in default_config[key]:
                config.set(key, subkey, default_config[key][subkey])

        with open(filename, "w", encoding="utf-8") as f:
            config.write(f)

        return config
    else:
        raise FileExistsError(
            f"Unable to write configuration file to {filename} because it already exists!"
        )


def verify_config(config: ConfigParser):
    default_config = get_default_config()
    for key in default_config:
        if not config.has_section(key):
            raise RuntimeError(f"Section {key} is required in configuration file.")
        for subkey in default_config[key]:
            if not config.has_option(key, subkey):
                raise RuntimeError(
                    f"Option {subkey} under section {key} is required in configuration file."
                )


def get_config(filename) -> ConfigParser:
    # If the config file doesn't exist, create it!
    if not os.path.isfile(filename):
        return create_default_config(filename)

    config = ConfigParser()
    with open(filename, "r", encoding="utf-8") as f:
        config.read_file(f)
    try:
        verify_config(config)
        return config
    except RuntimeError as e:
        print(f"[ERROR] Unable to start program: {repr(e)}")


if __name__ == "__main__":
    try:

        default_config_filename = "upcoming_drops_config.ini"

        filename = default_config_filename

        config = get_config(filename)

        logging.basicConfig(
            level=config.get("debug", "log_level").upper(),
            format="%(levelname)-7s : %(message)s",
        )

        filename = config.get("file_info", "filename")
        warning_title = config.get("appearance", "warning_title")
        warning_subtitle = config.get("appearance", "warning_subtitle")
        days = config.getint("functionality", "days_to_export")
        add_sheets_for_days = config.getboolean(
            "functionality", "additional_days_add_sheets"
        )

        um = UpcomingDrops(
            filename, warning_title, warning_subtitle, add_sheets_for_days
        )

        um.create_excel(days)
        input("Done! Press Enter to Exit...")
    except Exception as e:
        print("A fatal error has occurred: ", repr(e))
        input("Press ENTER to exit...")
        raise SystemExit
